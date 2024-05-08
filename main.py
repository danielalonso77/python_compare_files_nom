import os
from openpyxl import Workbook

# Crear un nuevo libro de Excel
wb = Workbook()
hoja = wb.active

# Agregar la cabecera en la primera fila de la primera columna
hoja.cell(row=1, column=1, value='Nomenclatura')

# Ruta de la carpeta que quieres explorar
ruta_carpeta = 'H:\OCTUBRE\quincena2319'

# Obtener la lista de archivos en la carpeta
archivos = os.listdir(ruta_carpeta)

# Escribir los nombres de los archivos debajo de la cabecera
for i, archivo in enumerate(archivos, start=2):
    # Dividir el nombre del archivo y su extensión
    nombre, extension = os.path.splitext(archivo)
    # Escribir el nombre del archivo sin extensión en la hoja de cálculo
    hoja.cell(row=i, column=1, value=nombre)

# Guardar el libro de Excel
wb.save('nombres_archivos.xlsx')
print("listo")