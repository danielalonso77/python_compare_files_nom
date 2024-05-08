from openpyxl import load_workbook, Workbook

# Cargar el archivo de Excel
archivo_entrada = 'H:\OCTUBRE\Indice de octubre del 2023.xlsx'
wb_entrada = load_workbook(archivo_entrada)
hoja_entrada = wb_entrada.active

# Crear un nuevo libro de Excel
wb_salida = Workbook()
hoja_salida = wb_salida.active

# Definir el número que se desea omitir
numero_a_omitir = 20  # Puedes cambiar este número según sea necesario

# Iterar sobre las filas y extraer las columnas A, C, F y G
for fila_entrada in hoja_entrada.iter_rows(min_row=2, values_only=True):  # Empezar desde la segunda fila
    columna_a = fila_entrada[0]  # Columna A

    # Verificar si el valor en la columna A coincide con el número a omitir
    if columna_a != numero_a_omitir:
        columna_c = fila_entrada[2]  # Columna C
        columna_f = fila_entrada[5]  # Columna F
        columna_g = fila_entrada[6]  # Columna G

        # Escribir los datos en el archivo de salida
        hoja_salida.append([columna_a, columna_c, columna_f, columna_g])

# Guardar el libro de Excel de salida
archivo_salida = 'nombre_del_archivo_salida.xlsx'
wb_salida.save(archivo_salida)
print('listo')

